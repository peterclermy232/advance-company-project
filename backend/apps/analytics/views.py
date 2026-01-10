from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Count, Q, F, Avg, Max
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from io import BytesIO

from apps.accounts.models import User
from apps.financial.models import FinancialAccount, Deposit


class AdminAnalyticsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    
    # Important: This tells DRF we don't need a detail route
    lookup_value_regex = '[^/.]+'
    
    def _check_admin(self, request):
        """Check if user is admin"""
        if request.user.role != 'admin':
            return Response(
                {'error': 'Admin access required'},
                status=status.HTTP_403_FORBIDDEN
            )
        return None
    
    @action(detail=False, methods=['get'], url_path='members')
    def members(self, request):
        """Get comprehensive member analytics"""
        admin_check = self._check_admin(request)
        if admin_check:
            return admin_check
        
        # Get all users with their financial accounts
        users = User.objects.filter(role='user').select_related('financial_account')
        
        # Calculate total contributions across all members
        total_all_contributions = FinancialAccount.objects.aggregate(
            total=Sum('total_contributions')
        )['total'] or Decimal('0')
        
        member_analytics = []
        
        for user in users:
            # Get or create financial account
            account, _ = FinancialAccount.objects.get_or_create(user=user)
            
            # Get deposit statistics
            deposit_stats = Deposit.objects.filter(
                user=user,
                status='completed'
            ).aggregate(
                total_deposits=Count('id'),
                last_deposit=Max('created_at')
            )
            
            # Calculate contribution percentage
            contribution_percentage = 0
            if total_all_contributions > 0:
                contribution_percentage = float(
                    (account.total_contributions / total_all_contributions) * 100
                )
            
            member_analytics.append({
                'id': user.id,
                'full_name': user.full_name,
                'email': user.email,
                'phone_number': user.phone_number,
                'total_contributions': float(account.total_contributions),
                'total_deposits': deposit_stats['total_deposits'],
                'interest_earned': float(account.interest_earned),
                'activity_status': user.activity_status,
                'created_at': user.created_at.isoformat(),
                'last_deposit_date': deposit_stats['last_deposit'].isoformat() if deposit_stats['last_deposit'] else None,
                'contribution_percentage': round(contribution_percentage, 2)
            })
        
        # Calculate summary statistics
        summary = {
            'total_contributions': float(total_all_contributions),
            'total_members': users.count(),
            'active_members': users.filter(activity_status='Active').count(),
            'average_contribution': float(total_all_contributions / users.count()) if users.count() > 0 else 0,
            'total_deposits_count': Deposit.objects.filter(status='completed').count(),
            'total_interest_earned': float(
                FinancialAccount.objects.aggregate(
                    total=Sum('interest_earned')
                )['total'] or Decimal('0')
            ),
        }
        
        # Get monthly trends (last 12 months)
        monthly_trends = []
        for i in range(11, -1, -1):
            month_start = datetime.now().replace(day=1) - timedelta(days=30 * i)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_data = Deposit.objects.filter(
                status='completed',
                created_at__gte=month_start,
                created_at__lte=month_end
            ).aggregate(
                total_amount=Sum('amount'),
                deposit_count=Count('id')
            )
            
            monthly_trends.append({
                'month': month_start.strftime('%b'),
                'year': month_start.year,
                'total_amount': float(month_data['total_amount'] or 0),
                'deposit_count': month_data['deposit_count'],
            })
        
        return Response({
            'members': member_analytics,
            'summary': summary,
            'monthly_trends': monthly_trends
        })
    
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Get analytics summary only"""
        admin_check = self._check_admin(request)
        if admin_check:
            return admin_check
        
        users = User.objects.filter(role='user')
        total_contributions = FinancialAccount.objects.aggregate(
            total=Sum('total_contributions')
        )['total'] or Decimal('0')
        
        summary = {
            'total_contributions': float(total_contributions),
            'total_members': users.count(),
            'active_members': users.filter(activity_status='Active').count(),
            'average_contribution': float(total_contributions / users.count()) if users.count() > 0 else 0,
            'total_deposits_count': Deposit.objects.filter(status='completed').count(),
            'total_interest_earned': float(
                FinancialAccount.objects.aggregate(
                    total=Sum('interest_earned')
                )['total'] or Decimal('0')
            ),
        }
        
        return Response(summary)
    
    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export analytics to Excel or PDF"""
        admin_check = self._check_admin(request)
        if admin_check:
            return admin_check
        
        export_format = request.query_params.get('format', 'excel')
        
        if export_format not in ['excel', 'pdf']:
            return Response(
                {'error': 'Invalid format. Use "excel" or "pdf"'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get member data
        users = User.objects.filter(role='user').select_related('financial_account')
        
        member_data = []
        for user in users:
            account, _ = FinancialAccount.objects.get_or_create(user=user)
            deposit_count = Deposit.objects.filter(user=user, status='completed').count()
            
            member_data.append({
                'name': user.full_name,
                'email': user.email,
                'contributions': float(account.total_contributions),
                'deposits': deposit_count,
                'interest': float(account.interest_earned),
                'status': user.activity_status
            })
        
        if export_format == 'excel':
            return self._export_excel(member_data)
        else:  # pdf
            return self._export_pdf(member_data)
    
    def _export_excel(self, member_data):
        """Export to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Member Analytics"
        
        # Header styling
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ['Name', 'Email', 'Total Contributions', 'Deposits', 'Interest Earned', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, member in enumerate(member_data, start=2):
            ws.cell(row=row, column=1, value=member['name'])
            ws.cell(row=row, column=2, value=member['email'])
            ws.cell(row=row, column=3, value=f"KES {member['contributions']:,.2f}")
            ws.cell(row=row, column=4, value=member['deposits'])
            ws.cell(row=row, column=5, value=f"KES {member['interest']:,.2f}")
            ws.cell(row=row, column=6, value=member['status'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="member-analytics-{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    def _export_pdf(self, member_data):
        """Export to PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("Member Analytics Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Date
        date_text = Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        )
        elements.append(date_text)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Summary statistics
        total_contributions = sum(m['contributions'] for m in member_data)
        total_interest = sum(m['interest'] for m in member_data)
        
        summary_text = Paragraph(
            f"<b>Summary:</b> Total Members: {len(member_data)} | "
            f"Total Contributions: KES {total_contributions:,.2f} | "
            f"Total Interest: KES {total_interest:,.2f}",
            styles['Normal']
        )
        elements.append(summary_text)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Table data
        table_data = [['Name', 'Email', 'Contributions', 'Deposits', 'Interest', 'Status']]
        
        for member in member_data:
            table_data.append([
                member['name'],
                member['email'],
                f"KES {member['contributions']:,.2f}",
                str(member['deposits']),
                f"KES {member['interest']:,.2f}",
                member['status']
            ])
        
        # Create table
        table = Table(table_data, colWidths=[1.5*inch, 2*inch, 1.3*inch, 0.8*inch, 1.3*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="member-analytics-{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        return response